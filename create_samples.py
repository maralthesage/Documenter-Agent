"""
Create Sample Excel Files for Testing
Generates realistic Excel files to test the documentation agent
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import random
import os


def create_sales_data_excel():
    """Create a sample sales data Excel file"""

    # Create data
    data = {
        'Date': pd.date_range('2024-01-01', periods=50, freq='D'),
        'Product_ID': [f'PROD_{random.randint(100, 999)}' for _ in range(50)],
        'Product_Name': [random.choice(['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Headphones']) for _ in range(50)],
        'Quantity_Sold': [random.randint(1, 20) for _ in range(50)],
        'Unit_Price': [random.choice([29.99, 49.99, 99.99, 199.99, 299.99]) for _ in range(50)],
        'Discount_Percent': [random.choice([0, 5, 10, 15, 20]) for _ in range(50)],
        'Total_Revenue': [0] * 50,  # Will be calculated
        'Region': [random.choice(['North', 'South', 'East', 'West']) for _ in range(50)],
        'Salesperson': [random.choice(['John Smith', 'Jane Doe', 'Mike Johnson', 'Sarah Williams']) for _ in range(50)],
    }

    df = pd.DataFrame(data)

    # Calculate Total_Revenue: (Quantity_Sold * Unit_Price) * (1 - Discount_Percent/100)
    df['Total_Revenue'] = (df['Quantity_Sold'] * df['Unit_Price'] * (1 - df['Discount_Percent'] / 100)).round(2)

    # Save to Excel
    output_file = './sample_data/sales_data.xlsx'
    os.makedirs('./sample_data', exist_ok=True)

    df.to_excel(output_file, index=False, sheet_name='Sales')

    # Add formatting using openpyxl
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Define colors
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    currency_format = '[$-409]#,##0.00;-[$-409]#,##0.00'
    date_format = 'MM/DD/YYYY'

    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Format columns
    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1):
        for col_idx, cell in enumerate(row, 1):
            if col_idx == 1:  # Date
                cell.number_format = date_format
            elif col_idx in [5, 7]:  # Unit_Price, Total_Revenue
                cell.number_format = currency_format

    # Highlight high revenue rows
    revenue_col_idx = 7
    for row_idx in range(2, len(df) + 2):
        cell = ws[f'G{row_idx}']
        try:
            value = float(cell.value) if cell.value else 0
            if value > 3000:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                cell.font = Font(color='006100')
        except:
            pass

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 15

    wb.save(output_file)
    print(f"✓ Created: {output_file}")
    return output_file


def create_inventory_excel():
    """Create a sample inventory Excel file"""

    data = {
        'Item_Code': [f'SKU-{random.randint(10000, 99999)}' for _ in range(40)],
        'Item_Name': [random.choice(['Widget A', 'Widget B', 'Gadget X', 'Gadget Y', 'Component Z']) for _ in range(40)],
        'Category': [random.choice(['Electronics', 'Hardware', 'Software', 'Supplies']) for _ in range(40)],
        'Current_Stock': [random.randint(0, 1000) for _ in range(40)],
        'Reorder_Level': [random.randint(50, 200) for _ in range(40)],
        'Unit_Cost': [round(random.uniform(10, 500), 2) for _ in range(40)],
        'Last_Updated': [datetime.now() - timedelta(days=random.randint(0, 30)) for _ in range(40)],
        'Supplier': [random.choice(['Supplier A', 'Supplier B', 'Supplier C']) for _ in range(40)],
        'Status': [''] * 40,  # Will be calculated
        'Stock_Value': [0] * 40,  # Will be calculated
    }

    df = pd.DataFrame(data)

    # Calculate Status: Low if Current_Stock < Reorder_Level, else Adequate
    df['Status'] = df.apply(
        lambda row: 'LOW' if row['Current_Stock'] < row['Reorder_Level'] else 'ADEQUATE',
        axis=1
    )

    # Calculate Stock_Value: Current_Stock * Unit_Cost
    df['Stock_Value'] = (df['Current_Stock'] * df['Unit_Cost']).round(2)

    output_file = './sample_data/inventory.xlsx'
    os.makedirs('./sample_data', exist_ok=True)

    df.to_excel(output_file, index=False, sheet_name='Inventory')

    # Add formatting
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Format columns
    currency_format = '[$-409]#,##0.00;-[$-409]#,##0.00'
    date_format = 'MM/DD/YYYY'

    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1):
        for col_idx, cell in enumerate(row, 1):
            if col_idx == 7:  # Last_Updated
                cell.number_format = date_format
            elif col_idx in [6, 10]:  # Unit_Cost, Stock_Value
                cell.number_format = currency_format

    # Highlight LOW status
    for row_idx in range(2, len(df) + 2):
        cell = ws[f'I{row_idx}']
        if cell.value == 'LOW':
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            cell.font = Font(color='9C0006')

    # Set column widths
    for col_idx, width in enumerate([12, 15, 15, 14, 14, 11, 13, 12, 10, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(output_file)
    print(f"✓ Created: {output_file}")
    return output_file


def create_financial_report_excel():
    """Create a sample financial report Excel file"""

    months = pd.date_range('2023-01-01', periods=12, freq='MS')
    data = {
        'Month': months,
        'Revenue': [random.randint(50000, 150000) for _ in range(12)],
        'COGS': [random.randint(20000, 60000) for _ in range(12)],
        'Gross_Profit': [0] * 12,
        'Operating_Expenses': [random.randint(10000, 30000) for _ in range(12)],
        'Operating_Income': [0] * 12,
        'Interest_Expense': [random.randint(1000, 5000) for _ in range(12)],
        'Income_Before_Tax': [0] * 12,
        'Tax_Rate': [0.25] * 12,
        'Net_Income': [0] * 12,
    }

    df = pd.DataFrame(data)

    # Calculate derived columns
    df['Gross_Profit'] = df['Revenue'] - df['COGS']
    df['Operating_Income'] = df['Gross_Profit'] - df['Operating_Expenses']
    df['Income_Before_Tax'] = df['Operating_Income'] - df['Interest_Expense']
    df['Net_Income'] = df['Income_Before_Tax'] * (1 - df['Tax_Rate'])

    output_file = './sample_data/financial_report.xlsx'
    os.makedirs('./sample_data', exist_ok=True)

    df.to_excel(output_file, index=False, sheet_name='P&L')

    # Add formatting
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Format columns
    currency_format = '[$-409]#,##0;-[$-409]#,##0'
    date_format = 'MMM YYYY'
    percent_format = '0.00%'

    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1):
        for col_idx, cell in enumerate(row, 1):
            if col_idx == 1:  # Month
                cell.number_format = date_format
            elif col_idx in range(2, 9):  # All currency columns
                cell.number_format = currency_format
            elif col_idx == 9:  # Tax_Rate
                cell.number_format = percent_format

    # Set column widths
    for col_idx, width in enumerate([12, 12, 12, 14, 18, 16, 16, 14, 10, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(output_file)
    print(f"✓ Created: {output_file}")
    return output_file


def main():
    """Create all sample files"""
    print("\n" + "=" * 60)
    print("Creating Sample Excel Files for Testing")
    print("=" * 60 + "\n")

    print("Creating sample data files...")
    files = [
        create_sales_data_excel(),
        create_inventory_excel(),
        create_financial_report_excel(),
    ]

    print("\n" + "=" * 60)
    print("Sample files created successfully!")
    print("=" * 60)
    print("\nYou can now test with:")
    for file in files:
        print(f"  python test.py '{file}'")

    return files


if __name__ == "__main__":
    main()
