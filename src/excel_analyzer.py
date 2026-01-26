"""
Excel Analysis Module
Handles parsing, analysis, and extraction of Excel file data including
formulas, formatting, data types, and structure.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.utils import get_column_letter
import re
from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass
import json


@dataclass
class ColumnInfo:
    """Data class for column information"""
    name: str
    index: int
    data_type: str
    non_null_count: int
    unique_values: int
    has_formula: bool
    formulas: List[str]
    sample_values: List[Any]
    min_value: Optional[Any] = None
    max_value: Optional[Any] = None
    fill_color: Optional[str] = None
    font_color: Optional[str] = None
    contains_link: bool = False
    link_type: Optional[str] = None
    numeric_range: Optional[Tuple[float, float]] = None
    
    def __post_init__(self):
        """Ensure all color values are strings or None"""
        # Convert fill_color to string if it's not None
        if self.fill_color is not None:
            self.fill_color = str(self.fill_color)
        
        # Convert font_color to string if it's not None
        if self.font_color is not None:
            self.font_color = str(self.font_color)


class ExcelAnalyzer:
    """Main class for analyzing Excel files"""

    def __init__(self, file_path: str, sheet_name: str = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.df = None
        self.workbook = None
        self.worksheet = None
        self.columns_info: List[ColumnInfo] = []

    def analyze(self) -> Dict[str, Any]:
        """
        Main analysis method that orchestrates all analysis tasks
        Returns a comprehensive analysis dictionary
        """
        print(f"Starting analysis of {self.file_path}")

        # Load data
        self._load_excel()

        # Analyze columns
        self._analyze_columns()

        # Extract formatting information
        self._extract_formatting()

        # Detect formulas
        self._detect_formulas()

        # Detect links and references
        self._detect_links()

        return self._generate_analysis_report()

    def _load_excel(self):
        """Load Excel file using both pandas and openpyxl"""
        # Get sheet names first
        xls = pd.ExcelFile(self.file_path)
        
        # Determine which sheet to use
        if self.sheet_name is None:
            # Use first sheet if not specified
            self.sheet_name = xls.sheet_names[0]
        elif self.sheet_name not in xls.sheet_names:
            # Try to find a matching sheet name (case-insensitive)
            matching_sheets = [s for s in xls.sheet_names if s.lower() == self.sheet_name.lower()]
            if matching_sheets:
                self.sheet_name = matching_sheets[0]
            else:
                raise ValueError(
                    f"Sheet '{self.sheet_name}' not found. "
                    f"Available sheets: {xls.sheet_names}"
                )
        
        print(f"Reading sheet: {self.sheet_name}")

        header_row = None
        # openpyxl for formatting and header detection
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            # Get the correct worksheet by name
            if self.sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[self.sheet_name]
            else:
                self.worksheet = self.workbook.active
            header_row = self._detect_header_row(self.worksheet)
        except Exception as e:
            print(f"Warning: Could not load workbook with openpyxl: {e}")
            # Fall back to active sheet for formula detection
            self.workbook = None
            self.worksheet = None

        # Pandas for data - explicitly specify sheet name
        if header_row is None:
            self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
        else:
            self.df = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_name,
                header=header_row - 1,
            )

        self._sanitize_column_names()

    def _detect_header_row(
        self,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        max_scan_rows: int = 20,
        min_non_empty: int = 2,
    ) -> Optional[int]:
        """Best-effort detection of the header row when it's not the first row."""
        best_row = None
        best_score = 0

        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), 1
        ):
            non_empty = [v for v in row if v not in (None, "")]
            if len(non_empty) < min_non_empty:
                continue

            string_like = [v for v in non_empty if isinstance(v, str) and v.strip()]
            if len(string_like) < max(1, len(non_empty) // 2):
                continue

            score = (len(string_like) * 2) + len(non_empty)
            if score > best_score:
                best_score = score
                best_row = row_idx

        return best_row

    def _sanitize_column_names(self) -> None:
        """Replace empty/unnamed columns with stable placeholders."""
        cleaned = []
        for idx, col in enumerate(self.df.columns, 1):
            if col is None or (isinstance(col, float) and pd.isna(col)):
                cleaned.append(f"Column {idx}")
                continue

            name = str(col).strip()
            if name.startswith("Unnamed"):
                cleaned.append(f"Column {idx}")
            else:
                cleaned.append(name)

        self.df.columns = cleaned

    def _analyze_columns(self):
        """Analyze each column in the dataframe"""
        for idx, col in enumerate(self.df.columns):
            col_data = self.df[col]

            # Determine data type
            data_type = self._infer_data_type(col_data)

            # Get statistics
            non_null = col_data.notna().sum()
            unique = col_data.nunique()

            # Get top 5 unique values (no repetition)
            sample_values = col_data.dropna().unique()[:5].tolist()

            # Get min/max for numeric columns
            min_val = None
            max_val = None
            numeric_range = None
            if data_type in ['int', 'float']:
                try:
                    min_val = col_data.min()
                    max_val = col_data.max()
                    numeric_range = (float(min_val), float(max_val))
                except:
                    pass

            col_info = ColumnInfo(
                name=str(col),
                index=idx,
                data_type=data_type,
                non_null_count=non_null,
                unique_values=unique,
                has_formula=False,  # Will be updated in _detect_formulas
                formulas=[],
                sample_values=sample_values,
                min_value=min_val,
                max_value=max_val,
                numeric_range=numeric_range,
            )

            self.columns_info.append(col_info)

    def _infer_data_type(self, series: pd.Series) -> str:
        """Infer the data type of a pandas series"""
        # Remove null values for analysis
        non_null = series.dropna()

        if len(non_null) == 0:
            return "empty"

        # Check for numeric types
        if pd.api.types.is_numeric_dtype(series):
            if pd.api.types.is_integer_dtype(series):
                return "int"
            else:
                return "float"

        # Check for datetime
        if pd.api.types.is_datetime64_any_dtype(series):
            return "datetime"

        # Check for boolean
        if pd.api.types.is_bool_dtype(series):
            return "boolean"

        # Check for actual boolean values in strings
        bool_vals = set(str(v).lower() for v in non_null)
        if bool_vals.issubset({'true', 'false', 'yes', 'no', '0', '1'}):
            return "categorical_boolean"

        # Check if all values are URLs
        if all(str(v).startswith(('http://', 'https://', 'ftp://')) for v in non_null):
            return "url"

        # Default to string/categorical
        return "string"

    def _detect_formulas(self):
        """Detect formulas in cells"""
        if self.worksheet is None:
            return  # Can't detect formulas without worksheet
        
        try:
            for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2, max_row=len(self.df) + 1), 1):
                for col_idx, cell in enumerate(row):
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # Found a formula
                        if col_idx - 1 < len(self.columns_info):
                            col_info = self.columns_info[col_idx - 1]
                            col_info.has_formula = True
                            if cell.value not in col_info.formulas:
                                col_info.formulas.append(cell.value)
        except Exception as e:
            print(f"Warning: Could not detect formulas: {e}")

    def _extract_formatting(self):
        """Extract formatting information (colors, styles)"""
        if self.worksheet is None:
            return  # Can't extract formatting without worksheet
        
        try:
            for col_idx, col_info in enumerate(self.columns_info, 1):
                col_letter = get_column_letter(col_idx)

                # Sample some cells to get formatting
                for row in range(2, min(10, len(self.df) + 2)):
                    try:
                        cell = self.worksheet[f'{col_letter}{row}']

                        # Extract fill color
                        if cell.fill and cell.fill.start_color:
                            try:
                                # Handle various color object types from openpyxl
                                color_obj = cell.fill.start_color
                                color = None
                                
                                # Try different attributes that might contain the color value
                                if hasattr(color_obj, 'rgb') and color_obj.rgb:
                                    color = str(color_obj.rgb)
                                elif hasattr(color_obj, 'index') and color_obj.index:
                                    color = str(color_obj.index)
                                elif hasattr(color_obj, 'theme') and color_obj.theme:
                                    color = str(color_obj.theme)
                                else:
                                    color = str(color_obj) if color_obj else None
                                
                                if color and color not in ['00000000', '00FFFFFF', 'None']:
                                    col_info.fill_color = str(color)  # Ensure string conversion
                                    break
                            except Exception as color_err:
                                pass

                        # Extract font color
                        if cell.font and cell.font.color:
                            try:
                                # Handle various color object types from openpyxl
                                color_obj = cell.font.color
                                color = None
                                
                                # Try different attributes that might contain the color value
                                if hasattr(color_obj, 'rgb') and color_obj.rgb:
                                    color = str(color_obj.rgb)
                                elif hasattr(color_obj, 'index') and color_obj.index:
                                    color = str(color_obj.index)
                                elif hasattr(color_obj, 'theme') and color_obj.theme:
                                    color = str(color_obj.theme)
                                else:
                                    color = str(color_obj) if color_obj else None
                                
                                if color and color not in ['00000000', '00FFFFFF', 'None']:
                                    col_info.font_color = str(color)  # Ensure string conversion
                            except Exception as color_err:
                                pass
                    except Exception as cell_err:
                        pass
        except Exception as e:
            print(f"Warning: Could not extract formatting: {e}")

    def _detect_links(self):
        """Detect VLOOKUP, INDEX/MATCH and other external references"""
        for col_info in self.columns_info:
            for formula in col_info.formulas:
                if 'VLOOKUP' in formula.upper():
                    col_info.contains_link = True
                    col_info.link_type = 'VLOOKUP'
                elif 'INDEX' in formula.upper() and 'MATCH' in formula.upper():
                    col_info.contains_link = True
                    col_info.link_type = 'INDEX_MATCH'
                elif formula.upper().startswith('=') and '!' in formula:
                    col_info.contains_link = True
                    col_info.link_type = 'EXTERNAL_REFERENCE'

    def _generate_analysis_report(self) -> Dict[str, Any]:
        """Generate comprehensive analysis report"""
        return {
            'file_path': self.file_path,
            'sheet_name': self.worksheet.title,
            'dimensions': {
                'rows': len(self.df),
                'columns': len(self.df.columns),
            },
            'columns': [
                {
                    'name': col.name,
                    'index': col.index,
                    'data_type': col.data_type,
                    'non_null_count': col.non_null_count,
                    'unique_values': col.unique_values,
                    'has_formula': col.has_formula,
                    'formulas': col.formulas,
                    'sample_values': [str(v) for v in col.sample_values],
                    'min_value': str(col.min_value) if col.min_value is not None else None,
                    'max_value': str(col.max_value) if col.max_value is not None else None,
                    'numeric_range': col.numeric_range,
                    'fill_color': col.fill_color,
                    'font_color': col.font_color,
                    'contains_link': col.contains_link,
                    'link_type': col.link_type,
                }
                for col in self.columns_info
            ],
        }

    def get_column_info(self) -> List[ColumnInfo]:
        """Return the list of analyzed columns"""
        return self.columns_info

    def extract_shared_formatting_groups(self) -> List[List[str]]:
        """Identify groups of columns with shared formatting"""
        groups = []
        color_groups = {}

        for col_info in self.columns_info:
            if col_info.fill_color:
                if col_info.fill_color not in color_groups:
                    color_groups[col_info.fill_color] = []
                color_groups[col_info.fill_color].append(col_info.name)

        groups = [cols for cols in color_groups.values() if len(cols) > 1]
        return groups
